"""Document parsing service for transaction extraction."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from ..schemas import ExtractedTransactionInput
from backend.shared.schemas.enums import TxnDirection


_DATE_FORMATS = (
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%d.%m.%Y",
    "%d-%b-%Y",
    "%d-%b-%y",
    "%d%b%Y",
    "%d%b%y",
)

_SUPPORTED_TYPES = {"csv", "xlsx", "txt", "pdf", "jpg", "jpeg", "png"}


class UnsupportedDocumentTypeError(ValueError):
    pass


@dataclass
class StatementContext:
    period_start: date | None = None
    period_end: date | None = None
    bank_code: str | None = None


@dataclass
class DocumentExtractionOutcome:
    """Result of parsing an uploaded file (rows + statement hints for persistence)."""

    rows: list[ExtractedTransactionInput]
    warnings: list[str]
    file_type: str
    ocr_pending: bool
    statement_context: StatementContext | None = None


def extract_transactions_from_document(
    *,
    filename: str,
    content_type: str | None,
    payload: bytes,
    bank_code_hint: str | None = None,
) -> DocumentExtractionOutcome:
    """Parse uploaded file and return normalized transaction rows."""
    file_type = _resolve_file_type(filename, content_type)
    if file_type not in _SUPPORTED_TYPES:
        raise UnsupportedDocumentTypeError(
            f"Unsupported file type '{file_type}'. Supported: {sorted(_SUPPORTED_TYPES)}",
        )

    if file_type in {"jpg", "jpeg", "png"}:
        return _parse_raster_image(payload, bank_code_hint, file_type)

    if file_type == "csv":
        rows, warnings = _parse_csv(payload, bank_code_hint)
        return DocumentExtractionOutcome(rows, warnings, file_type, False, None)

    if file_type == "xlsx":
        rows, warnings = _parse_xlsx(payload, bank_code_hint)
        return DocumentExtractionOutcome(rows, warnings, file_type, False, None)

    if file_type == "txt":
        rows, warnings = _parse_txt(payload, bank_code_hint)
        return DocumentExtractionOutcome(rows, warnings, file_type, False, None)

    return _parse_pdf(payload, bank_code_hint)


def _resolve_file_type(filename: str, content_type: str | None) -> str:
    ext = Path(filename).suffix.lower().lstrip(".")
    if ext:
        return ext
    if not content_type:
        return "unknown"
    if "csv" in content_type:
        return "csv"
    if "spreadsheet" in content_type or "excel" in content_type:
        return "xlsx"
    if "pdf" in content_type:
        return "pdf"
    if "text" in content_type:
        return "txt"
    if "jpeg" in content_type:
        return "jpeg"
    if "png" in content_type:
        return "png"
    return "unknown"


def _parse_csv(payload: bytes, bank_code_hint: str | None) -> tuple[list[ExtractedTransactionInput], list[str]]:
    text = payload.decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    rows: list[ExtractedTransactionInput] = []
    warnings: list[str] = []
    for idx, row in enumerate(reader, start=1):
        parsed = _parse_structured_row(idx, row, bank_code_hint)
        if parsed is None:
            warnings.append(f"Skipped CSV row {idx}: could not map date/description/amount.")
            continue
        rows.append(parsed)
    return rows, warnings


def _parse_xlsx(payload: bytes, bank_code_hint: str | None) -> tuple[list[ExtractedTransactionInput], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for xlsx parsing") from exc

    wb = load_workbook(io.BytesIO(payload), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return [], ["XLSX file is empty."]
    headers = [str(c).strip() if c is not None else "" for c in header]

    rows: list[ExtractedTransactionInput] = []
    warnings: list[str] = []
    for idx, values in enumerate(rows_iter, start=2):
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        parsed = _parse_structured_row(idx, row, bank_code_hint)
        if parsed is None:
            warnings.append(f"Skipped XLSX row {idx}: could not map date/description/amount.")
            continue
        rows.append(parsed)
    return rows, warnings


def _parse_txt(payload: bytes, bank_code_hint: str | None) -> tuple[list[ExtractedTransactionInput], list[str]]:
    lines = payload.decode("utf-8", errors="ignore").splitlines()
    ctx = _build_statement_context(lines, bank_code_hint)
    rows: list[ExtractedTransactionInput] = []
    warnings: list[str] = []
    for idx, line in enumerate(lines, start=1):
        parsed = _parse_free_text_line(idx, line, ctx)
        if parsed is None:
            continue
        rows.append(parsed)
    if not rows:
        warnings.append("No transaction-like lines detected in TXT file.")
    return rows, warnings


def _parse_pdf(payload: bytes, bank_code_hint: str | None) -> DocumentExtractionOutcome:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("pypdf is required for pdf parsing") from exc

    reader = PdfReader(io.BytesIO(payload))
    lines: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        lines.extend(text.splitlines())

    return _parse_statement_lines_from_text(lines, bank_code_hint, file_type="pdf")


def _parse_raster_image(
    payload: bytes,
    bank_code_hint: str | None,
    file_type: str,
) -> DocumentExtractionOutcome:
    lines, ocr_warnings = _ocr_image_bytes_to_lines(payload)
    if not lines:
        return DocumentExtractionOutcome(
            rows=[],
            warnings=ocr_warnings,
            file_type=file_type,
            ocr_pending=True,
            statement_context=None,
        )
    merged_lines = _merge_split_ocr_lines(lines)
    outcome = _parse_statement_lines_from_text(
        merged_lines,
        bank_code_hint,
        file_type=file_type,
        preamble_warnings=ocr_warnings,
    )
    return DocumentExtractionOutcome(
        outcome.rows,
        outcome.warnings,
        outcome.file_type,
        False,
        outcome.statement_context,
    )


def _parse_statement_lines_from_text(
    lines: list[str],
    bank_code_hint: str | None,
    *,
    file_type: str,
    preamble_warnings: list[str] | None = None,
) -> DocumentExtractionOutcome:
    context = _build_statement_context(lines, bank_code_hint)
    warnings: list[str] = list(preamble_warnings or [])

    if context.bank_code == "DIALOG_FINANCE":
        rows, w_df = _parse_dialog_finance_statement_lines(lines, context)
        warnings.extend(w_df)
        if rows:
            return DocumentExtractionOutcome(rows, warnings, file_type, False, context)

    if context.bank_code == "FRIMI":
        rows, w_fm = _parse_frimi_statement_lines(lines, context)
        warnings.extend(w_fm)
        if rows:
            return DocumentExtractionOutcome(rows, warnings, file_type, False, context)

    if context.bank_code == "NTB":
        rows, w_ntb = _parse_pdf_ntb(lines, context)
        warnings.extend(w_ntb)
        if rows:
            return DocumentExtractionOutcome(rows, warnings, file_type, False, context)

    if context.bank_code == "SAMPATH":
        rows, w_s = _parse_sampath_statement_lines(lines, context)
        warnings.extend(w_s)
        if rows:
            return DocumentExtractionOutcome(rows, warnings, file_type, False, context)

    rows: list[ExtractedTransactionInput] = []
    for idx, line in enumerate(lines, start=1):
        parsed = _parse_free_text_line(idx, line, context)
        if parsed:
            rows.append(parsed)
    if not rows:
        warnings.append(
            "No transaction rows detected in extracted text. "
            "For PDFs, the file may be image-only; for PNG/JPG, check Tesseract setup and scan quality.",
        )
    return DocumentExtractionOutcome(rows, warnings, file_type, False, context)


def _line_is_pure_compact_statement_date(s: str) -> bool:
    """True if *s* is only a DDMonYY / DDMonYYYY token (common table date column from OCR)."""
    t = re.sub(r"\s+", "", s.strip())
    return bool(re.fullmatch(r"\d{1,2}[A-Za-z]{2,5}\d{2,4}", t, re.IGNORECASE))


def _merge_split_ocr_lines(lines: list[str]) -> list[str]:
    """Join OCR lines when a statement date was read alone and amounts/description on the next line.

    Bank-agnostic: many statement screenshots put the date in one table cell and the rest of the
    row on the following OCR line (opening balance row, or narrow columns).
    """
    normalized = [re.sub(r"\s+", " ", ln).strip() for ln in lines]
    merged: list[str] = []
    i = 0
    while i < len(normalized):
        s = normalized[i]
        if not s:
            i += 1
            continue
        if _line_is_pure_compact_statement_date(s):
            j = i + 1
            while j < len(normalized) and not normalized[j]:
                j += 1
            if j < len(normalized):
                nxt = normalized[j]
                if (
                    not _line_is_pure_compact_statement_date(nxt)
                    and re.search(r"\d[\d,]*\.\d{2}", nxt)
                    and not re.match(
                        r"^\d{1,2}\s*[A-Za-z]{2,5}\s*\d{2,4}\s+\S",
                        nxt,
                        re.IGNORECASE,
                    )
                ):
                    merged.append(f"{s} {nxt}")
                    i = j + 1
                    continue
        merged.append(s)
        i += 1
    return merged


def _ocr_image_bytes_to_lines(payload: bytes) -> tuple[list[str], list[str]]:
    """Run Tesseract OCR; return text lines and human-readable warnings."""
    warnings: list[str] = []
    try:
        from PIL import Image
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Pillow is required for image OCR") from exc
    try:
        import pytesseract
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("pytesseract is required for image OCR") from exc

    try:
        from PIL import ImageOps

        img = Image.open(io.BytesIO(payload))
        img = img.convert("RGB")
        w, h = img.size
        if w < 1400:
            scale = 1400 / w
            img = img.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
        # Grayscale + contrast helps striped backgrounds and light watermarks (many bank PDFs).
        ocr_img = ImageOps.autocontrast(img.convert("L"), cutoff=1)
    except Exception as exc:
        return [], [f"Could not decode image: {exc}"]

    config = "--oem 3 --psm 6"
    try:
        text = pytesseract.image_to_string(ocr_img, config=config)
    except Exception as exc:
        err_name = type(exc).__name__
        msg = str(exc).lower()
        if "tesseract" in msg or err_name == "TesseractNotFoundError":
            warnings.append(
                "Tesseract OCR is not installed or not on PATH. "
                "On macOS: brew install tesseract. On Ubuntu: apt install tesseract-ocr.",
            )
        else:
            warnings.append(f"OCR failed ({err_name}): {exc}")
        return [], warnings

    if not text.strip():
        warnings.append("OCR returned no text; try a higher-resolution export.")
        return [], warnings

    return text.splitlines(), warnings


_MONTH_ABBREVS: tuple[str, ...] = (
    "jan",
    "feb",
    "mar",
    "apr",
    "may",
    "jun",
    "jul",
    "aug",
    "sep",
    "oct",
    "nov",
    "dec",
)

# Two-letter month tokens OCR often mangles (e.g. "Apr" → "Ar").
_OCR_MONTH_SHORT_FIX: dict[str, str] = {
    "ar": "apr",
    "ap": "apr",
    "mr": "mar",
    "ma": "mar",
    "my": "may",
    "jn": "jun",
    "ju": "jul",
    "jl": "jul",
    "au": "aug",
    "se": "sep",
    "oc": "oct",
    "no": "nov",
    "de": "dec",
    "ja": "jan",
    "fe": "feb",
}


def _normalize_sampath_compact_date(raw: str) -> str:
    """Normalize Sampath-style compact dates for :func:`_parse_date` (fixes OCR month typos)."""
    compact = re.sub(r"\s+", "", str(raw).strip())
    m = re.fullmatch(r"(\d{1,2})([A-Za-z]{2,5})(\d{2}|\d{4})", compact, re.IGNORECASE)
    if not m:
        return compact
    day_s, mon_raw, yr_s = m.group(1), m.group(2).lower(), m.group(3)
    mon3: str
    if len(mon_raw) >= 3 and mon_raw[:3] in _MONTH_ABBREVS:
        mon3 = mon_raw[:3]
    elif mon_raw in _OCR_MONTH_SHORT_FIX:
        mon3 = _OCR_MONTH_SHORT_FIX[mon_raw]
    else:
        hits = [a for a in _MONTH_ABBREVS if a.startswith(mon_raw) or mon_raw.startswith(a[:2])]
        mon3 = hits[0] if len(hits) == 1 else mon_raw[: min(3, len(mon_raw))]

    # %d%b%y / %d%b%Y expect Title-case month (Apr).
    return f"{day_s}{mon3.title()}{yr_s}"


def _parse_sampath_statement_lines(
    lines: list[str],
    context: StatementContext,
) -> tuple[list[ExtractedTransactionInput], list[str]]:
    """Parse Sampath-style statement rows (PDF text or OCR from PNG/JPG)."""
    rows: list[ExtractedTransactionInput] = []
    warnings: list[str] = []
    prev_balance: Decimal | None = None
    row_idx = 0

    header_starts = (
        "date ",
        "particulars",
        "debit",
        "credit",
        "balance",
        "statement of account",
        "account name",
        "account no",
        "branch",
        "www.",
        "page ",
    )

    for raw in lines:
        line = re.sub(r"\s+", " ", raw).strip()
        if len(line) < 8:
            continue
        low = line.lower()
        if any(low.startswith(p) for p in header_starts):
            continue
        if "statement period" in low:
            continue

        m = re.match(
            r"^(?P<d>\d{1,2}\s*[A-Za-z]{2,5}\s*\d{2,4})\s+(?P<rest>.+)$",
            line,
            re.IGNORECASE,
        )
        if not m:
            continue

        date_raw = re.sub(r"\s+", "", m.group("d"))
        date_for_parse = _normalize_sampath_compact_date(date_raw)
        rest = m.group("rest").replace("|", " ").strip()

        if re.match(r"^(c/f|b/f)\b", rest, re.IGNORECASE):
            continue

        nums = list(re.finditer(r"[-+]?\d[\d,]*\.\d{2}", rest))
        if len(nums) == 1:
            lone = _parse_decimal(nums[0].group(0))
            low_rest = rest.lower()
            head = rest[: nums[0].start()].strip(" -|")
            tail = rest[nums[0].end() :].strip(" -|").lower()
            is_opening_keywords = any(
                x in low_rest for x in ("balance", "b/f", "brought forward", "opening", "forward")
            )
            amount_only_tail = (not head) and (
                tail in {"", "cr", "dr", "cr.", "dr."}
                or tail.startswith("cr")
                or tail.startswith("dr")
            )
            if lone is not None and prev_balance is None and (
                is_opening_keywords or amount_only_tail
            ):
                prev_balance = lone
                tx_o = _parse_date(date_for_parse, context)
                if tx_o is not None:
                    desc_open = head or "Opening / brought forward balance"
                    row_idx += 1
                    rows.append(
                        ExtractedTransactionInput(
                            row_index=row_idx,
                            tx_date=tx_o.isoformat(),
                            raw_desc=desc_open,
                            amount_lkr=lone.quantize(Decimal("0.01")),
                            direction=TxnDirection.CR,
                            bank_code=context.bank_code or "SAMPATH",
                            parse_confidence=0.68,
                        ),
                    )
            continue
        if len(nums) < 2:
            continue

        amt_tok = nums[-2]
        bal_tok = nums[-1]
        amount = _parse_decimal(amt_tok.group(0))
        balance = _parse_decimal(bal_tok.group(0))
        if amount is None or balance is None:
            continue

        desc = rest[: amt_tok.start()].strip(" -|")
        if len(desc) < 2:
            continue
        low_desc = desc.lower()
        if "total debits" in low_desc or "total credits" in low_desc:
            continue
        if low_desc.startswith("total ") and ("debit" in low_desc or "credit" in low_desc):
            continue

        tx_date = _parse_date(date_for_parse, context)
        if tx_date is None:
            continue

        direction = _infer_direction_from_balance(prev_balance, balance, desc)
        prev_balance = balance

        row_idx += 1
        rows.append(
            ExtractedTransactionInput(
                row_index=row_idx,
                tx_date=tx_date.isoformat(),
                raw_desc=desc,
                amount_lkr=abs(amount).quantize(Decimal("0.01")),
                direction=direction,
                bank_code=context.bank_code or "SAMPATH",
                parse_confidence=0.72,
            ),
        )

    if not rows:
        warnings.append("Sampath layout parser found no rows (OCR noise or non-Sampath layout).")
    return rows, warnings


# pypdf often omits the space between closing balance and transaction date — allow \s* there.
_DIALOG_ROW_HEAD = re.compile(
    r"^([\d,]+\.\d{2})\s*(\d{1,2}-[A-Za-z]{3}-\d{4})\s+(\d{1,2}-[A-Za-z]{3}-\d{4})\s+(.*)$",
)
_DIALOG_FULL_ROW = re.compile(
    r"^([\d,]+\.\d{2})\s*(\d{1,2}-[A-Za-z]{3}-\d{4})\s+(\d{1,2}-[A-Za-z]{3}-\d{4})\s+"
    r"(.+?)\s+(\S+)\s+(\d[\d,]*\.\d{2})\s+(\d[\d,]*\.\d{2})\s*$",
)
_DIALOG_AMOUNTS = re.compile(r"^(\S+)\s+(\d[\d,]*\.\d{2})\s+(\d[\d,]*\.\d{2})\s*$")


def _dialog_finance_skip_meta_line(line: str) -> bool:
    s = line.strip()
    if not s:
        return True
    low = s.lower()
    if re.match(r"^--\s*\d+\s+of\s+\d+\s*--$", s):
        return True
    if low.startswith("consolidated individual monthly statement"):
        return True
    if "statement period" in low and "to" in low and re.search(r"\d{2}[-/]\d{2}[-/]\d{4}", s):
        return True
    if "www.dialogfinance.lk" in low:
        return True
    if "financialservice@dialog.lk" in low:
        return True
    if "weekdays" in low and "am" in low and "pm" in low:
        return True
    if re.match(r"^0\d{2}\s+\d\s+\d{3}\s+\d{3}", s):
        return True
    if "transaction date" in low and "value date" in low and "transaction details" in low:
        return True
    if "account number" in low and "opening balance" in low and "total deposits" in low:
        return True
    if low.startswith("savings account summary"):
        return True
    if low.startswith("individual account"):
        return True
    if re.match(r"^b/f\s+[\d,]+\.\d{2}$", low):
        return True
    if re.match(r"^total\s+[\d,]+\.\d{2}", low):
        return True
    if "total deposits" in low and "total withdrawals" in low:
        return True
    return False


def _dialog_finance_emit_row(
    row_idx: int,
    txn_date_raw: str,
    description: str,
    ref: str,
    deposit: Decimal,
    withdrawal: Decimal,
    context: StatementContext,
    warnings: list[str],
) -> tuple[int, ExtractedTransactionInput | None]:
    desc = re.sub(r"\s+", " ", description).strip()
    if not desc or re.match(r"^b/f\b", desc, re.IGNORECASE):
        return row_idx, None
    low = desc.lower()
    if low.startswith("total ") and "from_dfp" not in low:
        return row_idx, None

    tx_date = _parse_date(txn_date_raw.title(), context)
    if tx_date is None:
        return row_idx, None

    if deposit > 0 and withdrawal > 0:
        warnings.append(
            f"Dialog Finance: row has both deposit and withdrawal; using deposit: {desc[:60]}",
        )
        direction = TxnDirection.CR
        amount = deposit
    elif deposit > 0:
        direction = TxnDirection.CR
        amount = deposit
    elif withdrawal > 0:
        direction = TxnDirection.DR
        amount = withdrawal
    else:
        return row_idx, None

    if ref and ref not in desc:
        desc = f"{desc} ({ref})"

    row_idx += 1
    return row_idx, ExtractedTransactionInput(
        row_index=row_idx,
        tx_date=tx_date.isoformat(),
        raw_desc=desc,
        amount_lkr=amount.quantize(Decimal("0.01")),
        direction=direction,
        bank_code=context.bank_code or "DIALOG_FINANCE",
        parse_confidence=0.78,
    )


def _parse_dialog_finance_statement_lines(
    lines: list[str],
    context: StatementContext,
) -> tuple[list[ExtractedTransactionInput], list[str]]:
    """Parse Dialog Finance consolidated PDFs (balance column leads each row; details span lines)."""
    rows: list[ExtractedTransactionInput] = []
    warnings: list[str] = []
    row_idx = 0
    current: dict[str, object] | None = None

    i = 0
    n = len(lines)
    while i < n:
        line = re.sub(r"\s+", " ", lines[i]).strip()
        i += 1

        if line.startswith("Investments Summary"):
            if current:
                warnings.append(
                    "Dialog Finance: dropped incomplete row at investments section: "
                    f"{current.get('txn_date')!s}",
                )
                current = None
            break

        if _dialog_finance_skip_meta_line(line):
            continue

        if current is None:
            fm = _DIALOG_FULL_ROW.match(line)
            if fm:
                dep = _parse_decimal(fm.group(6)) or Decimal("0")
                wd = _parse_decimal(fm.group(7)) or Decimal("0")
                row_idx, emitted = _dialog_finance_emit_row(
                    row_idx,
                    fm.group(2),
                    fm.group(4).strip(),
                    fm.group(5),
                    dep,
                    wd,
                    context,
                    warnings,
                )
                if emitted:
                    rows.append(emitted)
                continue

            hm = _DIALOG_ROW_HEAD.match(line)
            if hm:
                tail = (hm.group(4) or "").strip()
                current = {
                    "txn_date": hm.group(2),
                    "value_date": hm.group(3),
                    "desc_parts": [tail] if tail else [],
                }
            continue

        am = _DIALOG_AMOUNTS.match(line)
        if am:
            dep = _parse_decimal(am.group(2)) or Decimal("0")
            wd = _parse_decimal(am.group(3)) or Decimal("0")
            parts = current.get("desc_parts") if current else None
            desc = " ".join(str(p) for p in parts) if isinstance(parts, list) else ""
            txn_date = str(current.get("txn_date")) if current else ""
            row_idx, emitted = _dialog_finance_emit_row(
                row_idx,
                txn_date,
                desc,
                am.group(1),
                dep,
                wd,
                context,
                warnings,
            )
            if emitted:
                rows.append(emitted)
            current = None
            continue

        hm = _DIALOG_ROW_HEAD.match(line)
        if hm and current:
            warnings.append(
                "Dialog Finance: incomplete row before new header; dropping partial description.",
            )
            current = None
            i -= 1
            continue

        if current:
            parts = current.get("desc_parts")
            if isinstance(parts, list):
                parts.append(line)

    if current:
        warnings.append("Dialog Finance: trailing incomplete row at EOF.")

    if not rows:
        warnings.append("Dialog Finance layout parser found no rows.")
    return rows, warnings


_FRIMI_LINE_DATE = re.compile(r"^\d{1,2}-[A-Za-z]{3}-\d{4}\s*$")


def _frimi_is_amount_token(s: str) -> bool:
    t = s.strip().replace(",", "")
    return bool(re.match(r"^\d+\.\d{2}$", t))


def _frimi_skip_footer_and_page_break(raw: list[str], idx: int, n: int) -> int:
    """Skip FriMi PDF footer blocks. pypdf often omits ``-- 1 of N --``; next page starts with ``Entry``."""
    while idx < n:
        low = raw[idx].strip().lower()
        if "frimi.lk" in low:
            idx += 1
            guard = 0
            while idx < n and guard < 40:
                guard += 1
                s = raw[idx].strip()
                if s == "Entry":
                    return idx
                if s.startswith("--") and (" of " in s or "end of statement" in s.lower()):
                    idx += 1
                    return idx
                idx += 1
            return idx
        s = raw[idx].strip()
        if s.startswith("--") and (" of " in s or "end of statement" in s.lower()):
            idx += 1
            continue
        break
    return idx


def _parse_frimi_statement_lines(
    lines: list[str],
    context: StatementContext,
) -> tuple[list[ExtractedTransactionInput], list[str]]:
    """Parse FriMi PDFs where pypdf emits one token per line (columnar layout lost)."""
    raw = [ln.rstrip("\n") for ln in lines]
    n = len(raw)
    rows: list[ExtractedTransactionInput] = []
    warnings: list[str] = []
    row_idx = 0
    i = 0

    while i < n - 2:
        i = _frimi_skip_footer_and_page_break(raw, i, n)
        if i >= n - 2:
            break

        t0, t1 = raw[i].strip(), raw[i + 1].strip()
        header_noise = {
            "entry",
            "date",
            "value",
            "transaction",
            "details",
            "references",
            "debits",
            "credits",
            "balance",
        }
        if t0.lower() in header_noise or t1.lower() in header_noise:
            i += 1
            continue

        if not (_FRIMI_LINE_DATE.match(t0) and _FRIMI_LINE_DATE.match(t1)):
            i += 1
            continue

        body_start = i + 2
        triple_at: int | None = None
        j = body_start
        while j < n - 2:
            tj = raw[j].strip()
            if _FRIMI_LINE_DATE.match(tj) and j + 1 < n and _FRIMI_LINE_DATE.match(raw[j + 1].strip()):
                break
            if _frimi_is_amount_token(raw[j]) and _frimi_is_amount_token(raw[j + 1]) and _frimi_is_amount_token(
                raw[j + 2],
            ):
                triple_at = j
                break
            j += 1

        if triple_at is None:
            if j < n and _FRIMI_LINE_DATE.match(raw[j].strip()) and j + 1 < n and _FRIMI_LINE_DATE.match(
                raw[j + 1].strip(),
            ):
                i = j
            else:
                i += 1
            continue

        ref_line = raw[triple_at - 1].strip() if triple_at > body_start else ""
        desc_parts = [raw[k].strip() for k in range(body_start, triple_at - 1) if raw[k].strip()]
        desc = " ".join(desc_parts)
        debit = _parse_decimal(raw[triple_at]) or Decimal("0")
        credit = _parse_decimal(raw[triple_at + 1]) or Decimal("0")

        i = triple_at + 3

        if not desc:
            continue
        low_d = desc.lower()
        if low_d.startswith("b/f") or "b/f balance" in low_d:
            continue

        tx_date = _parse_date(t0.title(), context)
        if tx_date is None:
            continue

        if credit > 0 and debit > 0:
            warnings.append(f"FriMi: both debit and credit non-zero; using credit: {desc[:50]}")
            direction = TxnDirection.CR
            amount = credit
        elif credit > 0:
            direction = TxnDirection.CR
            amount = credit
        elif debit > 0:
            direction = TxnDirection.DR
            amount = debit
        else:
            continue

        if ref_line and ref_line not in desc:
            desc = f"{desc} ({ref_line})"

        row_idx += 1
        rows.append(
            ExtractedTransactionInput(
                row_index=row_idx,
                tx_date=tx_date.isoformat(),
                raw_desc=desc,
                amount_lkr=amount.quantize(Decimal("0.01")),
                direction=direction,
                bank_code=context.bank_code or "FRIMI",
                parse_confidence=0.76,
            ),
        )

    if not rows:
        warnings.append("FriMi layout parser found no rows (check PDF text extraction).")
    return rows, warnings


def _parse_pdf_ntb(
    lines: list[str],
    context: StatementContext,
) -> tuple[list[ExtractedTransactionInput], list[str]]:
    stitched: list[tuple[int, str]] = []
    warnings: list[str] = []

    current_row_idx: int | None = None
    current_line = ""
    for idx, raw in enumerate(lines, start=1):
        line = re.sub(r"\s+", " ", raw).strip()
        if not line:
            continue
        if _looks_like_ntb_row_start(line):
            if current_row_idx is not None and current_line:
                stitched.append((current_row_idx, current_line.strip()))
            current_row_idx = idx
            current_line = line
            continue
        if (
            current_row_idx is not None
            and _count_amount_tokens(current_line) < 2
            and _looks_like_ntb_row_continuation(line)
        ):
            current_line += f" {line}"

    if current_row_idx is not None and current_line:
        stitched.append((current_row_idx, current_line.strip()))

    parsed_rows: list[ExtractedTransactionInput] = []
    prev_balance: Decimal | None = None
    for row_index, line in stitched:
        parsed = _parse_ntb_line(row_index, line, context, prev_balance)
        if parsed is None:
            continue
        parsed_rows.append(parsed[0])
        prev_balance = parsed[1]

    if not parsed_rows:
        warnings.append("NTB parser detected no rows; fell back to generic parser.")
    return parsed_rows, warnings


def _parse_structured_row(
    row_index: int,
    row: dict[str, object],
    bank_code_hint: str | None,
) -> ExtractedTransactionInput | None:
    normalized = {_normalize_key(k): v for k, v in row.items()}
    raw_desc = _first_value(
        normalized,
        ("raw_desc", "description", "particulars", "details", "narration", "transaction_details"),
    )
    raw_date = _first_value(
        normalized,
        ("tx_date", "date", "transaction_date", "value_date"),
    )
    amount_raw = _first_value(
        normalized,
        ("amount_lkr", "amount", "value"),
    )
    debit_raw = _first_value(normalized, ("debit", "debits", "withdrawal", "withdrawals"))
    credit_raw = _first_value(normalized, ("credit", "credits", "deposit", "deposits"))

    tx_date = _parse_date(raw_date)
    if tx_date is None or not raw_desc:
        return None

    direction: TxnDirection | None = None
    amount: Decimal | None = None

    if debit_raw is not None and _parse_decimal(debit_raw) not in (None, Decimal("0")):
        direction = TxnDirection.DR
        amount = _parse_decimal(debit_raw)
    elif credit_raw is not None and _parse_decimal(credit_raw) not in (None, Decimal("0")):
        direction = TxnDirection.CR
        amount = _parse_decimal(credit_raw)
    elif amount_raw is not None:
        parsed_amount = _parse_decimal(amount_raw)
        if parsed_amount is not None:
            direction = TxnDirection.CR if parsed_amount >= 0 else TxnDirection.DR
            amount = abs(parsed_amount)

    if direction is None or amount is None:
        direction = _infer_direction_from_text(str(raw_desc))
        amount = _parse_decimal(amount_raw) if amount_raw is not None else None
        if amount is not None:
            amount = abs(amount)

    if amount is None:
        return None

    return ExtractedTransactionInput(
        row_index=row_index,
        tx_date=tx_date.isoformat(),
        raw_desc=str(raw_desc).strip(),
        amount_lkr=amount.quantize(Decimal("0.01")),
        direction=direction,
        bank_code=bank_code_hint,
        parse_confidence=0.85,
    )


def _parse_free_text_line(
    row_index: int,
    line: str,
    context: StatementContext,
) -> ExtractedTransactionInput | None:
    # Ex: "19-Dec ... CEFTS Charges ... 025.00 1,095,701.27"
    compact = re.sub(r"\s+", " ", line).strip()
    if len(compact) < 16:
        return None

    date_match = re.match(
        r"^(?P<date>\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/][A-Za-z]{3}(?:[-/]\d{2,4})?|\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|\d{1,2}[A-Za-z]{3}\d{2,4})\s+",
        compact,
    )
    if not date_match:
        return None
    tx_date = _parse_date(date_match.group("date"), context)
    if tx_date is None:
        return None

    numbers = list(re.finditer(r"[-+]?\d[\d,]*\.\d{2}", compact))
    if not numbers:
        return None

    amount_text = numbers[-2].group(0) if len(numbers) >= 2 else numbers[-1].group(0)
    amount = _parse_decimal(amount_text)
    if amount is None:
        return None

    desc_start = date_match.end()
    desc_end = numbers[-2].start() if len(numbers) >= 2 else numbers[-1].start()
    description = compact[desc_start:desc_end].strip(" -")
    if not description:
        return None

    return ExtractedTransactionInput(
        row_index=row_index,
        tx_date=tx_date.isoformat(),
        raw_desc=description,
        amount_lkr=abs(amount).quantize(Decimal("0.01")),
        direction=_infer_direction_from_text(description),
        bank_code=context.bank_code,
        parse_confidence=0.65,
    )


def _parse_ntb_line(
    row_index: int,
    line: str,
    context: StatementContext,
    prev_balance: Decimal | None,
) -> tuple[ExtractedTransactionInput, Decimal] | None:
    m = re.match(r"^(?P<d1>\d{1,2}-[A-Za-z]{3})\s+(?P<d2>\d{1,2}-[A-Za-z]{3})\s+(?P<rest>.+)$", line)
    if not m:
        return None

    tx_date = _parse_date(m.group("d1"), context)
    if tx_date is None:
        return None

    rest = m.group("rest")
    nums = list(re.finditer(r"[-+]?\d[\d,]*\.\d{2}", rest))
    if len(nums) < 2:
        return None

    amount = _parse_decimal(nums[-2].group(0))
    balance = _parse_decimal(nums[-1].group(0))
    if amount is None or balance is None:
        return None

    desc = rest[: nums[-2].start()].strip(" -")
    if not desc or "total " in desc.lower():
        return None

    direction = _infer_direction_from_balance(prev_balance, balance, desc)

    return (
        ExtractedTransactionInput(
            row_index=row_index,
            tx_date=tx_date.isoformat(),
            raw_desc=desc,
            amount_lkr=abs(amount).quantize(Decimal("0.01")),
            direction=direction,
            bank_code=context.bank_code,
            parse_confidence=0.82,
        ),
        balance,
    )


def _normalize_key(key: str | None) -> str:
    if key is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", key.strip().lower()).strip("_")


def _first_value(row: dict[str, object], keys: tuple[str, ...]) -> object | None:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def _parse_date(value: object | None, context: StatementContext | None = None) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None

    if re.fullmatch(r"\d{1,2}-[A-Za-z]{3}", text):
        year = context.period_end.year if context and context.period_end else date.today().year
        text = f"{text}-{year}"

    if re.fullmatch(r"\d{1,2}[A-Za-z]{3}\d{4}", text, re.IGNORECASE):
        try:
            parsed = datetime.strptime(text.title(), "%d%b%Y").date()
            return _anchor_to_statement_period(parsed, context)
        except ValueError:
            pass

    if re.fullmatch(r"\d{1,2}[A-Za-z]{3}\d{2}", text, re.IGNORECASE):
        try:
            parsed = datetime.strptime(text.title(), "%d%b%y").date()
            return _anchor_to_statement_period(parsed, context)
        except ValueError:
            pass

    for fmt in _DATE_FORMATS:
        try:
            parsed = datetime.strptime(text, fmt).date()
            return _anchor_to_statement_period(parsed, context)
        except ValueError:
            continue
    return None


def _parse_decimal(value: object | None) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    text = text.replace("CR", "").replace("DR", "").strip()
    if text in {"", "-", "--"}:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _infer_direction_from_text(text: str) -> TxnDirection:
    t = text.lower()
    credit_signals = ("salary", "interest", "deposit", "credit", "refund", "int.pd")
    if any(token in t for token in credit_signals):
        return TxnDirection.CR
    return TxnDirection.DR


def _build_statement_context(lines: list[str], bank_code_hint: str | None) -> StatementContext:
    context = StatementContext(bank_code=(bank_code_hint.upper() if bank_code_hint else None))
    period_re = re.compile(
        r"Statement Period\s*:\s*(\d{2}[-/]\d{2}[-/]\d{4})\s*to\s*(\d{2}[-/]\d{2}[-/]\d{4})",
        re.IGNORECASE,
    )
    frimi_period_inline = re.compile(
        r"(\d{1,2}-[A-Za-z]{3}-\d{4})\s+To\s+(\d{1,2}-[A-Za-z]{3}-\d{4})",
        re.IGNORECASE,
    )
    for idx, raw in enumerate(lines):
        line = raw.strip()
        if context.bank_code is None:
            low = line.lower()
            if "frimi.lk" in low:
                context.bank_code = "FRIMI"
            elif "dialogfinance.lk" in low or "financialservice@dialog.lk" in low:
                context.bank_code = "DIALOG_FINANCE"
            elif "nations trust bank" in low:
                context.bank_code = "NTB"
            elif "sampathbank" in low or "sampath bank" in low:
                context.bank_code = "SAMPATH"
        m = period_re.search(line)
        if m:
            context.period_start = datetime.strptime(m.group(1), "%d-%m-%Y").date()
            context.period_end = datetime.strptime(m.group(2), "%d-%m-%Y").date()
        fm = frimi_period_inline.search(line)
        if fm:
            try:
                context.period_start = datetime.strptime(fm.group(1), "%d-%b-%Y").date()
                context.period_end = datetime.strptime(fm.group(2), "%d-%b-%Y").date()
            except ValueError:
                pass
        if (
            context.period_start is None
            and _FRIMI_LINE_DATE.match(line)
            and idx + 2 < len(lines)
            and lines[idx + 1].strip().lower() == "to"
            and _FRIMI_LINE_DATE.match(lines[idx + 2].strip())
        ):
            try:
                context.period_start = datetime.strptime(line, "%d-%b-%Y").date()
                context.period_end = datetime.strptime(lines[idx + 2].strip(), "%d-%b-%Y").date()
            except ValueError:
                pass
    return context


def _anchor_to_statement_period(parsed: date, context: StatementContext | None) -> date:
    if context is None or context.period_start is None or context.period_end is None:
        return parsed
    anchored = parsed
    if anchored < context.period_start:
        # Interest/tax adjustments are often posted immediately after period end.
        anchored = anchored.replace(year=anchored.year + 1)
    if anchored > context.period_end and (anchored - context.period_end).days > 35:
        anchored = anchored.replace(year=anchored.year - 1)
    return anchored


def _looks_like_ntb_row_start(line: str) -> bool:
    if re.match(r"^\d{1,2}-[A-Za-z]{3}\s+\d{1,2}-[A-Za-z]{3}\s+", line):
        return True
    return False


def _looks_like_ntb_row_continuation(line: str) -> bool:
    lower = line.lower()
    if lower.startswith(("total ", "transaction", "statement period", "-- ")):
        return False
    if re.match(r"^\d{1,2}-[A-Za-z]{3}\s+\d{1,2}-[A-Za-z]{3}\s+", line):
        return False
    return bool(re.search(r"\d[\d,]*\.\d{2}", line))


def _count_amount_tokens(text: str) -> int:
    return len(re.findall(r"[-+]?\d[\d,]*\.\d{2}", text))


def _infer_direction_from_balance(
    prev_balance: Decimal | None,
    balance: Decimal,
    description: str,
) -> TxnDirection:
    if prev_balance is not None:
        if balance > prev_balance:
            return TxnDirection.CR
        if balance < prev_balance:
            return TxnDirection.DR
    return _infer_direction_from_text(description)
